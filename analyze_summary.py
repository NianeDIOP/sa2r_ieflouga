import openpyxl
import sys

# Force UTF-8 encoding for output
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

# Charger le fichier Excel
wb = openpyxl.load_workbook('c:/Users/niana/Downloads/SA2R-IEFLOUGA/base.xlsx', data_only=True)

print("=" * 100)
print("RÉSUMÉ DE L'ANALYSE COMPLÈTE DE BASE.XLSX")
print("=" * 100)
print()

# Analyser tous les onglets
sheets_info = []

for sheet_name in ['2-Effectifs', '1-Gene&Finan', '3-CI&CMG&CFEE', '4-Personnel', '5-Mat&Mob']:
    try:
        ws = wb[sheet_name]
        
        # Compter les colonnes avec données
        max_col_with_data = 0
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=3, column=col).value:  # Si ligne 3 (en-tête colonnes) a une valeur
                max_col_with_data = col
        
        sheets_info.append({
            'name': sheet_name,
            'total_cols': max_col_with_data
        })
        
        print(f"✓ {sheet_name}: {max_col_with_data} colonnes")
    except Exception as e:
        print(f"✗ {sheet_name}: Erreur - {e}")

print()
print("=" * 100)
print("TOTAL GÉNÉRAL")
print("=" * 100)

total_columns = sum(s['total_cols'] for s in sheets_info)
print(f"\nNOMBRE TOTAL DE COLONNES DANS LE FICHIER EXCEL: {total_columns}")
print()

for sheet in sheets_info:
    print(f"  - {sheet['name']}: {sheet['total_cols']} colonnes")

print()
print("=" * 100)
