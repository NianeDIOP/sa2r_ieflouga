import openpyxl
from collections import defaultdict

# Charger le fichier Excel
wb = openpyxl.load_workbook('c:/Users/niana/Downloads/SA2R-IEFLOUGA/base.xlsx', data_only=True)

# Sélectionner l'onglet "4-Personnel"
ws = wb['4-Personnel']

print("=" * 100)
print("ANALYSE DE L'ONGLET: 4-Personnel (ENSEIGNANTS ET PERSONNEL)")
print("=" * 100)
print()

# Analyser les 3 premières lignes d'en-tête
print("STRUCTURE DES EN-TÊTES:")
print("-" * 100)

# Ligne 1: Sections principales
print("\nLIGNE 1 (Sections principales):")
row1 = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value:
        row1.append((col, str(cell_value).strip()))
        print(f"  Col {col}: {cell_value}")

# Ligne 2: Sous-sections
print("\nLIGNE 2 (Sous-sections):")
row2 = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=2, column=col).value
    if cell_value:
        row2.append((col, str(cell_value).strip()))
        print(f"  Col {col}: {cell_value}")

# Ligne 3: Noms des colonnes
print("\nLIGNE 3 (Noms des colonnes):")
row3 = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=3, column=col).value
    if cell_value:
        row3.append((col, str(cell_value).strip()))
        print(f"  Col {col}: {cell_value}")

# Compter le nombre total de colonnes
print("\n" + "=" * 100)
print(f"NOMBRE TOTAL DE COLONNES: {ws.max_column}")
print("=" * 100)

# Organiser les colonnes par sections
print("\nORGANISATION HIÉRARCHIQUE:")
print("-" * 100)

# Créer une structure hiérarchique
current_section = None
current_subsection = None
sections = defaultdict(lambda: defaultdict(list))

for col in range(1, ws.max_column + 1):
    v1 = ws.cell(row=1, column=col).value
    v2 = ws.cell(row=2, column=col).value
    v3 = ws.cell(row=3, column=col).value
    
    if v1:
        current_section = str(v1).strip()
    if v2:
        current_subsection = str(v2).strip()
    
    if v3:
        column_name = str(v3).strip()
        sections[current_section][current_subsection].append((col, column_name))

# Afficher la structure organisée
for section, subsections in sections.items():
    print(f"\n📂 {section}")
    for subsection, columns in subsections.items():
        print(f"  📁 {subsection if subsection != 'None' else '(Sans sous-section)'}")
        for col_num, col_name in columns:
            print(f"    └─ Col {col_num:3d}: {col_name}")

# Résumé par section
print("\n" + "=" * 100)
print("RÉSUMÉ PAR SECTION:")
print("=" * 100)
for section, subsections in sections.items():
    total_cols = sum(len(columns) for columns in subsections.values())
    print(f"{section}: {total_cols} colonnes")
    for subsection, columns in subsections.items():
        print(f"  - {subsection if subsection != 'None' else '(Sans sous-section)'}: {len(columns)} colonnes")

print("\n" + "=" * 100)
print(f"TOTAL GÉNÉRAL: {ws.max_column} COLONNES")
print("=" * 100)
