import openpyxl

# Charger le fichier Excel
wb = openpyxl.load_workbook('base.xlsx')

print('='*80)
print('ANALYSE DETAILLEE DU FICHIER base.xlsx')
print('='*80)

for i, sheet_name in enumerate(wb.sheetnames):
    print(f'\n\n[ONGLET {i+1}: {sheet_name}]')
    print('-'*80)
    
    ws = wb[sheet_name]
    
    # Récupérer les en-têtes (première ligne)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f'Nombre de colonnes: {len(headers)}')
    print(f'Nombre de lignes de données: {ws.max_row - 1}')
    print(f'\nCOLONNES:')
    
    for j, header in enumerate(headers):
        print(f'  {j+1:3d}. {header}')
    
    # Afficher un exemple de données (ligne 2)
    if ws.max_row > 1:
        print(f'\nEXEMPLE DE DONNEES (ligne 2):')
        row_data = []
        for cell in ws[2]:
            if cell.value:
                row_data.append(str(cell.value)[:50])  # Limiter à 50 caractères
        if row_data:
            for j, (header, data) in enumerate(zip(headers, row_data)):
                if data:
                    print(f'  {header}: {data}')

print('\n' + '='*80)
print('ANALYSE TERMINEE')
print('='*80)
