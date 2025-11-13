import openpyxl
import json

# Charger le fichier Excel
wb = openpyxl.load_workbook('base.xlsx')

# Structure complète
structure = {}

# Analyser chaque onglet principal
for sheet_name in wb.sheetnames:
    if sheet_name in ['Ref', 'Param', 'Feuil2']:
        continue  # Skip reference sheets
    
    ws = wb[sheet_name]
    
    # Extraire les en-têtes (lignes 1, 2, 3)
    headers_row1 = [cell.value for cell in ws[1]]
    headers_row2 = [cell.value for cell in ws[2]]
    headers_row3 = [cell.value for cell in ws[3]] if ws.max_row >= 3 else []
    
    # Organiser les colonnes par section
    sections = {}
    current_section = None
    current_subsection = None
    
    for col_idx in range(len(headers_row2)):
        h1 = headers_row1[col_idx] if col_idx < len(headers_row1) else None
        h2 = headers_row2[col_idx]
        h3 = headers_row3[col_idx] if col_idx < len(headers_row3) else None
        
        if h2:  # Si la ligne 2 a une valeur
            # Déterminer la section (ligne 1)
            if h1 and h1 != current_section:
                current_section = h1
                if current_section not in sections:
                    sections[current_section] = []
            
            # Si pas de section définie, utiliser "GENERAL"
            if current_section is None:
                current_section = "GENERAL"
                if current_section not in sections:
                    sections[current_section] = []
            
            # Ajouter la colonne
            sections[current_section].append({
                'column': chr(65 + col_idx) if col_idx < 26 else f'{chr(65 + col_idx//26 - 1)}{chr(65 + col_idx%26)}',
                'name': h2,
                'index': col_idx
            })
    
    structure[sheet_name] = sections

# Sauvegarder en JSON
with open('excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(structure, f, indent=2, ensure_ascii=False)

print("Structure sauvegardée dans excel_structure.json")

# Afficher un résumé
print("\n" + "="*100)
print("RÉSUMÉ DE LA STRUCTURE")
print("="*100)

for sheet_name, sections in structure.items():
    print(f'\n\n[{sheet_name}]')
    print(f'Nombre de sections: {len(sections)}')
    total_cols = sum(len(cols) for cols in sections.values())
    print(f'Nombre total de colonnes: {total_cols}')
    
    for section_name, columns in sections.items():
        print(f'\n  Section: {section_name}')
        print(f'  - {len(columns)} colonnes')
        # Afficher les 5 premières colonnes
        for col in columns[:5]:
            print(f'    • {col["name"]}')
        if len(columns) > 5:
            print(f'    ... et {len(columns) - 5} autres')
