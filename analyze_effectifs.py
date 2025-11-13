import openpyxl
from openpyxl.utils import get_column_letter

# Charger le fichier Excel
wb = openpyxl.load_workbook('base.xlsx')

# Analyser l'onglet Effectifs
ws = wb['2-Effectifs']

print('='*100)
print('ANALYSE DÉTAILLÉE DE L\'ONGLET: 2-Effectifs')
print('='*100)
print(f'\nNombre total de colonnes: {ws.max_column}')
print(f'Nombre de lignes: {ws.max_row}')

# Afficher toutes les en-têtes des 3 premières lignes
print('\n' + '='*100)
print('STRUCTURE COMPLÈTE DES EN-TÊTES')
print('='*100)

# Ligne 1 (Sections principales)
print('\n--- LIGNE 1 (Sections principales) ---')
row1_sections = {}
current_section = None
for col_num in range(1, ws.max_column + 1):
    cell = ws.cell(1, col_num)
    col_letter = get_column_letter(col_num)
    if cell.value:
        current_section = cell.value
        print(f'{col_letter}: {cell.value}')
        row1_sections[col_num] = cell.value

# Ligne 2 (Sous-sections)
print('\n--- LIGNE 2 (Sous-sections/Niveaux) ---')
row2_headers = {}
for col_num in range(1, ws.max_column + 1):
    cell = ws.cell(2, col_num)
    col_letter = get_column_letter(col_num)
    if cell.value:
        row2_headers[col_num] = cell.value
        section = None
        # Trouver la section parente
        for sec_col, sec_name in sorted(row1_sections.items(), reverse=True):
            if col_num >= sec_col:
                section = sec_name
                break
        print(f'{col_letter}: [{section}] -> {cell.value}')

# Ligne 3 (Détails des colonnes)
print('\n--- LIGNE 3 (Colonnes détaillées) ---')
print('\nTOUTES LES COLONNES PAR SECTION:')
print('='*100)

current_section = None
section_columns = {}

for col_num in range(1, ws.max_column + 1):
    cell1 = ws.cell(1, col_num)
    cell2 = ws.cell(2, col_num)
    cell3 = ws.cell(3, col_num)
    col_letter = get_column_letter(col_num)
    
    # Déterminer la section
    if cell1.value:
        current_section = cell1.value
        if current_section not in section_columns:
            section_columns[current_section] = []
    
    # Ajouter la colonne
    if cell3.value:
        section_columns[current_section].append({
            'col': col_letter,
            'num': col_num,
            'section': cell1.value if cell1.value else '',
            'subsection': cell2.value if cell2.value else '',
            'name': cell3.value
        })

# Afficher par section
for section, columns in section_columns.items():
    print(f'\n{"="*100}')
    print(f'SECTION: {section}')
    print(f'Nombre de colonnes: {len(columns)}')
    print(f'{"="*100}')
    
    for col in columns:
        subsection = f" [{col['subsection']}]" if col['subsection'] else ""
        print(f"  {col['col']:3s} (col {col['num']:3d}): {col['name']}{subsection}")

# Afficher quelques exemples de données
print('\n' + '='*100)
print('EXEMPLES DE DONNÉES (Ligne 4)')
print('='*100)

for col_num in range(1, min(50, ws.max_column + 1)):
    cell3 = ws.cell(3, col_num)
    cell4 = ws.cell(4, col_num)
    col_letter = get_column_letter(col_num)
    
    if cell3.value and cell4.value:
        print(f'{col_letter}: {cell3.value} = {cell4.value}')

print('\n' + '='*100)
print('ANALYSE TERMINÉE')
print('='*100)
