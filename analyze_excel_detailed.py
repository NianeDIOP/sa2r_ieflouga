import openpyxl
from openpyxl.utils import get_column_letter

# Charger le fichier Excel
wb = openpyxl.load_workbook('base.xlsx')

print('='*100)
print('ANALYSE DETAILLEE COMPLETE DU FICHIER base.xlsx')
print('='*100)

for i, sheet_name in enumerate(wb.sheetnames):
    print(f'\n\n{"="*100}')
    print(f'ONGLET {i+1}: {sheet_name}')
    print(f'{"="*100}')
    
    ws = wb[sheet_name]
    
    print(f'Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes')
    
    # Analyser les 5 premières lignes pour comprendre la structure
    print(f'\n--- STRUCTURE DES 5 PREMIERES LIGNES ---')
    
    for row_num in range(1, min(6, ws.max_row + 1)):
        print(f'\nLigne {row_num}:')
        row_data = []
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row_num, col_num)
            if cell.value:
                col_letter = get_column_letter(col_num)
                row_data.append(f'{col_letter}: {str(cell.value)[:60]}')
        
        if row_data:
            for item in row_data:
                print(f'  {item}')
        else:
            print('  (vide)')

print('\n' + '='*100)
print('ANALYSE TERMINEE')
print('='*100)
